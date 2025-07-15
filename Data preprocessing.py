import os
import pandas as pd
import numpy as np
from pykalman import KalmanFilter
from scipy import stats
import warnings
import logging
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from scipy import interpolate
import xgboost as xgb
from sklearn.model_selection import train_test_split
from sklearn.metrics import mean_squared_error

# 设置日志
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# 忽略 SettingWithCopyWarning
warnings.filterwarnings("ignore", category=pd.errors.SettingWithCopyWarning)

def load_data(file_path):
    """Load CSV data and convert timestamp to datetime."""
    df = pd.read_csv(file_path)
    
    # 检查时间列的名称
    time_column = 'ts_interval_x' if 'ts_interval_x' in df.columns else 'ts_interval_ref'
    
    df[time_column] = pd.to_datetime(df[time_column])
    df.set_index(time_column, inplace=True)

    # 如果存在PSP列，删除它
    if 'PSP' in df.columns:
        print(f"文件 {os.path.basename(file_path)} 中删除了PSP列")
        df = df.drop(columns=['PSP'])

    return df

def apply_temperature_filter(data):
    # 定义合理的温度范围
    absolute_min = -30
    absolute_max = 30

    # 计算每小时的温度变化
    temp_diff = np.abs(np.diff(data))
    
    # 定义异常的温度变化阈值（比如5度/小时）
    max_hourly_change = 5

    # 初始化卡尔曼滤波器
    kf = KalmanFilter(initial_state_mean=0, n_dim_obs=1)
    filtered_state_means, _ = kf.filter(data)

    # 创建一个掩码数组，初始全为False
    anomaly_mask = np.zeros(len(data), dtype=bool)

    # 标记异常值
    anomaly_mask[0] = data[0] < absolute_min or data[0] > absolute_max
    for i in range(1, len(data)):
        if (data[i] < absolute_min or data[i] > absolute_max or
            temp_diff[i-1] > max_hourly_change):
            anomaly_mask[i] = True

    # 只替换被标记为异常的值
    filtered_data = np.where(anomaly_mask, filtered_state_means.flatten(), data)

    return filtered_data, anomaly_mask

def apply_dswrf_filter(data):
    # 初始化卡尔曼滤波器
    kf = KalmanFilter(initial_state_mean=0, n_dim_obs=1)
    
    # 应用卡尔曼滤波
    filtered_state_means, _ = kf.filter(data)
    
    # 对于太阳辐射，我们只需要处理异常的高值
    upper_bound = np.percentile(data[data > 0], 99.9)  # 使用99.9百分位数作为上界
    lower_bound = 0  # 下界保持为0

    # 异常检测和填充
    anomaly_mask = (data < lower_bound) | (data > upper_bound)
    filtered_data = np.where(anomaly_mask, filtered_state_means.flatten(), data)
    
    return filtered_data, anomaly_mask

def apply_3sigma_criterion(df, column, window_size='144h', step_size='24h'):
    """Apply 3-sigma criterion for a specific column."""
    result = df[column].copy()
    modified = pd.Series(False, index=df.index)
    for start in pd.date_range(df.index.min(), df.index.max() - pd.Timedelta(window_size), freq=step_size):
        end = start + pd.Timedelta(window_size)
        window = df.loc[start:end, column].dropna()
        
        if len(window) > 1:
            mean = window.mean()
            std = window.std()
            if std != 0:
                outliers = np.abs(window - mean) > 3 * std
                result.loc[window.index[outliers]] = np.nan
                modified.loc[window.index[outliers]] = True
    
    return result, modified

def detect_anomalies(df, column, threshold=0.04):
    """Detect anomalies based on rate of change for a specific column."""
    result = df[column].copy()
    modified = pd.Series(False, index=df.index)
    rate_of_change = result.pct_change().abs()
    
    for i in range(len(rate_of_change)):
        if rate_of_change.iloc[i] > threshold:
            # 检查前后3个时刻的变化率
            start = max(0, i-3)
            end = min(len(rate_of_change), i+4)
            surrounding_rates = rate_of_change.iloc[start:end]
            surrounding_rates = surrounding_rates[surrounding_rates.index != rate_of_change.index[i]]
            
            # 如果所有周围的变化率都正常，则认为当前点是正常的
            if all(surrounding_rates <= threshold):
                continue
            
            result.iloc[i] = np.nan
            modified.iloc[i] = True
    
    return result, modified

def detect_consecutive_zeros(series, threshold=1):
    """检测连续的0值"""
    return (series == 0).rolling(window=threshold, min_periods=threshold).sum() >= threshold

def detect_large_deviations(series, threshold=0.5):
    """检测与平均值相差超过指定阈值的数据"""
    mean = series.mean()
    return abs((series - mean) / mean) > threshold

def process_column(df, column, window_size='7D', step_size='1D', anomaly_threshold=0.04):
    """Process a single column with 3-sigma criterion and anomaly detection."""
    # 应用3-sigma准则
    df[column], modified_3sigma = apply_3sigma_criterion(df, column, window_size, step_size)
    
    # 检测基于阈值的异常值
    df[column], modified_anomalies = detect_anomalies(df, column, anomaly_threshold)
    
    modified = modified_3sigma | modified_anomalies
    
    # 对PIH列进行特殊处理
    if column == 'PIH':
        # 创建一个布尔掩码，标识超出0-100范围的值
        pih_outliers = (df[column] < 0) | (df[column] > 100)
        # 将超出范围的值设为NaN
        df.loc[pih_outliers, column] = np.nan
        # 更新修改标记
        modified = modified | pih_outliers
    
    # 对于其他特定列，检测连续的0值
    if column in ['PIF', 'PIH', 'water_temp', 'PSP', 'PRT', 'PRP', 'PCH', 'SST', 'SSP', 'SRT', 'SRP']:
        consecutive_zeros = detect_consecutive_zeros(df[column])
        df.loc[consecutive_zeros, column] = np.nan
        modified_zeros = pd.Series(False, index=df.index)
        modified_zeros[consecutive_zeros] = True
        modified = modified | modified_zeros
    
    return df, modified

def cubic_spline_interpolate(series):
    """对给定的序列进行三次样条插值"""
    non_null = series.notnull()
    indices = np.arange(len(series))
    x = indices[non_null]
    y = series[non_null].values
    
    if len(x) > 3:  # 三次样条插值至少需要4个点
        cs = interpolate.CubicSpline(x, y)
        return pd.Series(cs(indices), index=series.index)
    else:
        return series  # 如果点数不足，返回原始序列

def xgboost_predict(df, target_column, gap):
    """使用XGBoost预测缺失值"""
    # 准备特征和目标变量
    features = df.columns.drop(target_column).tolist()
    X = df[features]
    y = df[target_column]
    
    # 分割训练集和测试集
    X_train, X_test, y_train, y_test = train_test_split(X[y.notnull()], y[y.notnull()], test_size=0.3, random_state=42)
    
    # 创建和训练XGBoost模型
    model = xgb.XGBRegressor(random_state=42)
    model.fit(X_train, y_train)
    
    # 预测缺失值
    X_missing = X.iloc[gap]
    predictions = model.predict(X_missing)
    
    # 将预测值填充到原始序列中
    filled_series = y.copy()
    filled_series.iloc[gap] = predictions
    
    return filled_series

def fill_missing_values(df):
    """填充缺失值，不限制填充值的变化率"""
    original_df = df.copy()
    result = df.copy()
    modified = pd.DataFrame(False, index=df.index, columns=df.columns)
    
    for column in df.columns:
        # 找到所有缺失值的位置
        null_mask = df[column].isnull()
        null_indices = np.where(null_mask)[0]
        
        # 如果没有缺失值，继续下一列
        if len(null_indices) == 0:
            continue
        
        # 分割连续缺失值的区间
        gaps = np.split(null_indices, np.where(np.diff(null_indices) != 1)[0] + 1)
        
        for gap in gaps:
            if len(gap) <= 3:
                # 使用三次样条插值
                interpolated = cubic_spline_interpolate(result[column])
            else:
                # 使用XGBoost预测
                interpolated = xgboost_predict(result, column, gap)
            
            # 直接填充插值或预测的值，不进行变化率限制
            result[column].iloc[gap] = interpolated.iloc[gap]
        
        # 对于特定列，将填充值的最小值设为0，不能出现负数
        if column in ['water_temp', 'PSP', 'PRT', 'PRP', 'PIF', 'PIH', 'PCH', 'PREVOS', 'SST', 'SSP', 'SRT', 'SRP']:
            result[column] = result[column].clip(lower=0)
        
        # 更新修改标记
        modified[column] = result[column] != original_df[column]
    
    # 使用前向填充和后向填充处理剩余的 NaN 值
    result = result.fillna(method='ffill').fillna(method='bfill')
    
    logging.info(f"After filling missing values: {result.notna().sum().sum()} non-null values")
    return result, modified

def save_to_excel(df, modified, output_file):
    """Save the processed dataframe to an Excel file with color-coded cells."""
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df.to_excel(writer, index=True, sheet_name='Processed Data')
        
        workbook = writer.book
        worksheet = writer.sheets['Processed Data']
        
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        
        for col in range(len(df.columns) + 1):  # +1 for index column
            for row in range(len(df) + 1):  # +1 for header row
                if row == 0 or col == 0:  # Skip header row and index column
                    continue
                if modified.iloc[row-1, col-1]:
                    cell = worksheet.cell(row=row+1, column=col+1)
                    cell.fill = yellow_fill
    
    logging.info(f"Saved processed data to Excel file: {output_file}")

def process_file(input_file, output_file, column_settings=None):
    print(f"正在处理文件: {os.path.basename(input_file)}")
    
    try:
        # 加载数据
        df = load_data(input_file)
        logging.info(f"Loaded data shape: {df.shape}, non-null values: {df.notna().sum().sum()}")
        
        modified = pd.DataFrame(False, index=df.index, columns=df.columns)
        
        # 处理 item_temperature 列
        temp_data = df['item_temperature'].values
        temp_filtered, temp_modified = apply_temperature_filter(temp_data)
        df['item_temperature'] = temp_filtered
        modified['item_temperature'] = temp_modified
        
        # 处理 item_dswrf 列
        dswrf_data = df['item_dswrf'].values
        dswrf_filtered, dswrf_modified = apply_dswrf_filter(dswrf_data)
        df['item_dswrf'] = dswrf_filtered
        modified['item_dswrf'] = dswrf_modified
        
        # 处理其他列
        if column_settings is None:
            column_settings = {}
        
        for column in df.columns:
            if column not in ['item_temperature', 'item_dswrf']:
                settings = column_settings.get(column, {})
                window_size = settings.get('window_size', '7D')
                step_size = settings.get('step_size', '1D')
                anomaly_threshold = settings.get('anomaly_threshold', 0.04)
                
                df, column_modified = process_column(df, column, window_size, step_size, anomaly_threshold)
                modified[column] = column_modified
        
        # 填补缺失值
        df, fill_modified = fill_missing_values(df)
        modified = modified | fill_modified
        
        # 保存处理后的数据到Excel文件
        save_to_excel(df, modified, output_file)
        print(f"处理后的数据已保存到 {output_file}")
    except Exception as e:
        print(f"处理文件时出错：{str(e)}")
        print(f"错误发生在以下行：")
        import traceback
        traceback.print_exc()

def main(input_dir, output_dir, column_settings=None):
    # 获取绝对路径
    input_dir = os.path.abspath(input_dir)
    output_dir = os.path.abspath(output_dir)

    print(f"使用的输入目录绝对路径: {input_dir}")
    print(f"使用的输出目录绝对路径: {output_dir}")

    # 验证输入目录是否存在
    if not os.path.exists(input_dir):
        print(f"错误：输入目录 '{input_dir}' 不存在。")
        return

    # 验证输入路径是否是一个目录
    if not os.path.isdir(input_dir):
        print(f"错误：'{input_dir}' 不是一个目录。")
        return

    # 创建输出目录（如果不存在）
    os.makedirs(output_dir, exist_ok=True)

    failed_files = []  # 保存失败文件的列表
    
    # 处理目录中的所有CSV文件
    for filename in os.listdir(input_dir):
        if filename.endswith('.csv'):
            input_file = os.path.join(input_dir, filename)
            output_file = os.path.join(output_dir, f"cleaned_{filename.replace('.csv', '.xlsx')}")
            
            # 如果处理失败，记录文件名
            if not process_file(input_file, output_file, column_settings):
                failed_files.append(filename)
    

if __name__ == "__main__":
    input_dir = r"E:\PythonProject\converted_data"
    output_dir = r"E:\PythonProject\cleanded_data"

    #需要自行调整参数
    column_settings = {
        'water_temp': {'window_size': '7D', 'step_size': '1D', 'anomaly_threshold': 0.01},
        'PSP': {'window_size': '7D', 'step_size': '1D', 'anomaly_threshold': 0.01},
        'PRT': {'window_size': '7D', 'step_size': '1D', 'anomaly_threshold': 0.01},
        'PRP': {'window_size': '7D', 'step_size': '1D', 'anomaly_threshold': 0.01},
        'PIF': {'window_size': '7D', 'step_size': '1D', 'anomaly_threshold': 0.01},
        'PIH': {'window_size': '7D', 'step_size': '1D', 'anomaly_threshold': 0.01},
        'PCH': {'window_size': '7D', 'step_size': '1D', 'anomaly_threshold': 0.01},
        'PREVOS': {'window_size': '7D', 'step_size': '1D', 'anomaly_threshold': 0.01},
        'SST': {'window_size': '7D', 'step_size': '1D', 'anomaly_threshold': 0.01},
        'SSP': {'window_size': '7D', 'step_size': '1D', 'anomaly_threshold': 0.01},
        'SRT': {'window_size': '7D', 'step_size': '1D', 'anomaly_threshold': 0.01},
        'SRP': {'window_size': '7D', 'step_size': '1D', 'anomaly_threshold': 0.01},
        'HNCHS': {'window_size': '7D', 'step_size': '1D', 'anomaly_threshold': 0.01},
        'HNISHS': {'window_size': '7D', 'step_size': '1D', 'anomaly_threshold': 0.01},
        'HSOP': {'window_size': '7D', 'step_size': '1D', 'anomaly_threshold': 0.01},
        'HSOT': {'window_size': '7D', 'step_size': '1D', 'anomaly_threshold': 0.01},
        'HSOF': {'window_size': '7D', 'step_size': '1D', 'anomaly_threshold': 0.01},
        'HSRWP': {'window_size': '7D', 'step_size': '1D', 'anomaly_threshold': 0.01},
        'HSRWT': {'window_size': '7D', 'step_size': '1D', 'anomaly_threshold': 0.01},
        'HSRWFL': {'window_size': '7D', 'step_size': '1D', 'anomaly_threshold': 0.01},
        'item_humidity': {'window_size': '7D', 'step_size': '1D', 'anomaly_threshold': 0.01},
        'item_wbTemperature': {'window_size': '7D', 'step_size': '1D', 'anomaly_threshold': 0.01},
        'item_enthalpy': {'window_size': '7D', 'step_size': '1D', 'anomaly_threshold': 0.01},
        'item_cloudRate': {'window_size': '7D', 'step_size': '1D', 'anomaly_threshold': 0.01},
        'item_windSpeed': {'window_size': '7D', 'step_size': '1D', 'anomaly_threshold': 0.01},
        'item_windDirection': {'window_size': '7D', 'step_size': '1D', 'anomaly_threshold': 0.01},
        'item_pressure': {'window_size': '7D', 'step_size': '1D', 'anomaly_threshold': 0.01},

 }



    print(f"输入目录: {input_dir}")
    print(f"输出目录: {output_dir}")
    
    main(input_dir, output_dir, column_settings)